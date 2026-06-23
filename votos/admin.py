import unicodedata

from django import forms
from django.contrib import admin, messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl import Workbook, load_workbook

# Register your models here.
from .models import Casilla, Municipio, Partido, ResultadoCasilla, Seccion


class ImportarSeccionesForm(forms.Form):
    archivo = forms.FileField(
        label="Archivo Excel",
        help_text="Usa un archivo .xlsx con columnas municipio y seccion.",
    )

    def clean_archivo(self):
        archivo = self.cleaned_data["archivo"]

        if not archivo.name.lower().endswith(".xlsx"):
            raise forms.ValidationError("El archivo debe estar en formato .xlsx.")

        return archivo


class ImportarCasillasForm(forms.Form):
    municipio = forms.ModelChoiceField(
        queryset=Municipio.objects.all(),
        label="Municipio",
        help_text="Las casillas del Excel se importaran dentro de este municipio.",
    )
    archivo = forms.FileField(
        label="Archivo Excel",
        help_text=(
            "Usa un archivo .xlsx con columnas seccion, tipo y numero."
        ),
    )

    def clean_archivo(self):
        archivo = self.cleaned_data["archivo"]

        if not archivo.name.lower().endswith(".xlsx"):
            raise forms.ValidationError("El archivo debe estar en formato .xlsx.")

        return archivo


def normalizar_encabezado(valor):
    texto = str(valor or "").strip().lower()
    texto = "".join(
        caracter
        for caracter in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caracter)
    )
    return " ".join(texto.replace("_", " ").split())


def buscar_columna(encabezados, *opciones):
    for opcion in opciones:
        columna = encabezados.get(opcion)
        if columna is not None:
            return columna

    return None


def normalizar_tipo_casilla(valor):
    texto = normalizar_encabezado(valor)
    tipos = {
        "b": Casilla.TIPO_BASICA,
        "basica": Casilla.TIPO_BASICA,
        "basico": Casilla.TIPO_BASICA,
        "c": Casilla.TIPO_CONTIGUA,
        "contigua": Casilla.TIPO_CONTIGUA,
        "e": Casilla.TIPO_EXTRAORDINARIA,
        "extraordinaria": Casilla.TIPO_EXTRAORDINARIA,
        "s": Casilla.TIPO_ESPECIAL,
        "especial": Casilla.TIPO_ESPECIAL,
    }
    return tipos.get(texto)


def crear_excel_resultados(queryset):
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resultados"
    hoja.append(
        [
            "Municipio",
            "Seccion",
            "Tipo casilla",
            "Numero casilla",
            "Partido",
            "Nombre partido",
            "Tipo partido",
            "Votos",
            "Total acta",
            "Total calculado",
            "Diferencia",
            "Tiene diferencia",
            "Usuario captura",
            "Fecha captura",
        ]
    )

    for resultado in queryset.select_related(
        "casilla",
        "casilla__seccion",
        "casilla__seccion__municipio",
        "casilla__usuario_captura",
        "partido",
    ):
        casilla = resultado.casilla
        seccion = casilla.seccion
        partido = resultado.partido
        hoja.append(
            [
                seccion.municipio.nombre,
                seccion.numero,
                casilla.get_tipo_display(),
                casilla.numero,
                partido.siglas,
                partido.nombre,
                partido.get_tipo_display(),
                resultado.votos,
                casilla.total_acta,
                casilla.total_calculado,
                casilla.diferencia,
                "Si" if casilla.tiene_diferencia else "No",
                casilla.usuario_captura.username if casilla.usuario_captura else "",
                casilla.fecha_captura,
            ]
        )

    for columna in hoja.columns:
        ancho = max(len(str(celda.value or "")) for celda in columna)
        hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 2, 40)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        'attachment; filename="resultados-casilla.xlsx"'
    )
    libro.save(response)
    return response


@admin.register(Municipio)
class MunicipioAdmin(admin.ModelAdmin):
    list_display = ["nombre"]
    search_fields = ["nombre"]


@admin.register(Seccion)
class SeccionAdmin(admin.ModelAdmin):
    change_list_template = "admin/votos/seccion/change_list.html"
    list_display = ["numero", "municipio"]
    list_filter = ["municipio"]
    search_fields = ["numero", "municipio__nombre"]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "importar-excel/",
                self.admin_site.admin_view(self.importar_excel),
                name="votos_seccion_importar_excel",
            ),
        ]
        return custom_urls + urls

    def importar_excel(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied

        if request.method == "POST":
            form = ImportarSeccionesForm(request.POST, request.FILES)

            if form.is_valid():
                archivo = form.cleaned_data["archivo"]

                try:
                    libro = load_workbook(archivo, read_only=True, data_only=True)
                    hoja = libro.active
                except Exception:
                    messages.error(request, "No se pudo leer el archivo Excel.")
                    return redirect("admin:votos_seccion_changelist")

                encabezados = {
                    normalizar_encabezado(celda.value): indice
                    for indice, celda in enumerate(hoja[1])
                }
                municipio_columna = encabezados.get("municipio")
                seccion_columna = buscar_columna(
                    encabezados,
                    "seccion",
                    "numero",
                    "numero seccion",
                )

                if municipio_columna is None or seccion_columna is None:
                    messages.error(
                        request,
                        "El Excel debe tener las columnas municipio y seccion.",
                    )
                    return redirect("admin:votos_seccion_changelist")

                creadas = 0
                existentes = 0
                omitidas = 0

                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    municipio_nombre = str(fila[municipio_columna] or "").strip()
                    seccion_valor = fila[seccion_columna]

                    if not municipio_nombre or not seccion_valor:
                        continue

                    try:
                        seccion_numero = int(seccion_valor)
                    except (TypeError, ValueError):
                        omitidas += 1
                        continue

                    municipio, _ = Municipio.objects.get_or_create(
                        nombre=municipio_nombre,
                    )
                    _, creada = Seccion.objects.get_or_create(
                        municipio=municipio,
                        numero=seccion_numero,
                    )

                    if creada:
                        creadas += 1
                    else:
                        existentes += 1

                messages.success(
                    request,
                    (
                        f"Importacion terminada. Secciones creadas: {creadas}. "
                        f"Ya existentes: {existentes}. Filas omitidas: {omitidas}."
                    ),
                )
                return redirect("admin:votos_seccion_changelist")
        else:
            form = ImportarSeccionesForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": "Importar secciones desde Excel",
            "opts": self.model._meta,
        }
        return render(request, "admin/votos/seccion/importar_excel.html", context)


@admin.register(Casilla)
class CasillaAdmin(admin.ModelAdmin):
    change_list_template = "admin/votos/casilla/change_list.html"
    list_display = [
        "seccion",
        "tipo",
        "numero",
        "total_acta",
        "total_calculado",
        "diferencia",
        "tiene_diferencia",
        "usuario_captura",
        "fecha_captura",
    ]
    list_filter = ["tipo", "seccion__municipio", "tiene_diferencia"]
    search_fields = ["seccion__numero", "seccion__municipio__nombre"]
    readonly_fields = [
        "total_calculado",
        "diferencia",
        "tiene_diferencia",
        "usuario_captura",
        "fecha_captura",
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "importar-excel/",
                self.admin_site.admin_view(self.importar_excel),
                name="votos_casilla_importar_excel",
            ),
        ]
        return custom_urls + urls

    def importar_excel(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied

        if request.method == "POST":
            form = ImportarCasillasForm(request.POST, request.FILES)

            if form.is_valid():
                archivo = form.cleaned_data["archivo"]

                try:
                    libro = load_workbook(archivo, read_only=True, data_only=True)
                    hoja = libro.active
                except Exception:
                    messages.error(request, "No se pudo leer el archivo Excel.")
                    return redirect("admin:votos_casilla_changelist")

                encabezados = {
                    normalizar_encabezado(celda.value): indice
                    for indice, celda in enumerate(hoja[1])
                }
                seccion_columna = buscar_columna(
                    encabezados,
                    "seccion",
                    "numero seccion",
                )
                tipo_columna = buscar_columna(encabezados, "tipo", "tipo casilla")
                numero_columna = buscar_columna(
                    encabezados,
                    "numero",
                    "numero casilla",
                    "casilla",
                )

                if (
                    seccion_columna is None
                    or tipo_columna is None
                    or numero_columna is None
                ):
                    messages.error(
                        request,
                        (
                            "El Excel debe tener las columnas seccion, "
                            "tipo y numero."
                        ),
                    )
                    return redirect("admin:votos_casilla_changelist")

                municipio = form.cleaned_data["municipio"]
                creadas = 0
                existentes = 0
                omitidas = 0

                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    seccion_valor = fila[seccion_columna]
                    tipo = normalizar_tipo_casilla(fila[tipo_columna])
                    numero_valor = fila[numero_columna]

                    if (
                        not seccion_valor
                        or not tipo
                        or not numero_valor
                    ):
                        omitidas += 1
                        continue

                    try:
                        seccion_numero = int(seccion_valor)
                        casilla_numero = int(numero_valor)
                    except (TypeError, ValueError):
                        omitidas += 1
                        continue

                    seccion, _ = Seccion.objects.get_or_create(
                        municipio=municipio,
                        numero=seccion_numero,
                    )
                    _, creada = Casilla.objects.get_or_create(
                        seccion=seccion,
                        tipo=tipo,
                        numero=casilla_numero,
                    )

                    if creada:
                        creadas += 1
                    else:
                        existentes += 1

                messages.success(
                    request,
                    (
                        f"Importacion terminada. Casillas creadas: {creadas}. "
                        f"Ya existentes: {existentes}. Filas omitidas: {omitidas}."
                    ),
                )
                return redirect("admin:votos_casilla_changelist")
        else:
            form = ImportarCasillasForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": "Importar casillas desde Excel",
            "opts": self.model._meta,
        }
        return render(request, "admin/votos/casilla/importar_excel.html", context)


@admin.register(Partido)
class PartidoAdmin(admin.ModelAdmin):
    list_display = ["orden_captura", "siglas", "nombre", "tipo", "color", "imagen_url"]
    list_editable = ["orden_captura"]
    list_display_links = ["siglas"]
    list_filter = ["tipo"]
    search_fields = ["siglas", "nombre"]


@admin.register(ResultadoCasilla)
class ResultadoCasillaAdmin(admin.ModelAdmin):
    change_list_template = "admin/votos/resultadocasilla/change_list.html"
    list_display = [
        "casilla",
        "partido",
        "votos",
        "total_acta",
        "total_calculado",
        "diferencia",
        "tiene_diferencia",
        "usuario_captura",
        "fecha_captura",
    ]
    list_filter = ["partido", "casilla__seccion__municipio", "casilla__tiene_diferencia"]
    search_fields = [
        "partido__siglas",
        "partido__nombre",
        "casilla__seccion__numero",
        "casilla__seccion__municipio__nombre",
    ]
    actions = ["exportar_seleccionados_excel"]

    @admin.display(description="Total acta")
    def total_acta(self, obj):
        return obj.casilla.total_acta

    @admin.display(description="Total calculado")
    def total_calculado(self, obj):
        return obj.casilla.total_calculado

    @admin.display(description="Diferencia")
    def diferencia(self, obj):
        return obj.casilla.diferencia

    @admin.display(boolean=True, description="Tiene diferencia")
    def tiene_diferencia(self, obj):
        return obj.casilla.tiene_diferencia

    @admin.display(description="Usuario captura")
    def usuario_captura(self, obj):
        return obj.casilla.usuario_captura

    @admin.display(description="Fecha captura")
    def fecha_captura(self, obj):
        return obj.casilla.fecha_captura

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "exportar-excel/",
                self.admin_site.admin_view(self.exportar_excel),
                name="votos_resultadocasilla_exportar_excel",
            ),
        ]
        return custom_urls + urls

    def exportar_excel(self, request):
        if not self.has_view_permission(request):
            raise PermissionDenied

        changelist = self.get_changelist_instance(request)
        queryset = changelist.get_queryset(request)
        return crear_excel_resultados(queryset)

    @admin.action(description="Exportar seleccionados a Excel")
    def exportar_seleccionados_excel(self, request, queryset):
        if not self.has_view_permission(request):
            raise PermissionDenied

        return crear_excel_resultados(queryset)
